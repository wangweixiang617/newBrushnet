import multiprocessing as mp
import re
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import baostock as bs
import pandas as pd
from openpyxl.styles import Font, PatternFill
from tqdm.auto import tqdm



# ========================== 标的配置 ==========================
# 支持二元组或三元组；程序只读取前两个字段：
#   (配置名称, Baostock代码)
# 第三个字段即使存在也不会用于最终名称。
# 最终显示名称和 Excel Sheet 名称优先使用 Baostock 远程返回的证券名称。
ETF_CONFIG1 = [
    ("科创50ETF", "sh.588000"),
    ("中证银行ETF", "sh.512800"),
    # ("黄金ETF", "sh.518880"),
    # ("恒生科技ETF", "sh.513180"),
    # ("港股通科技ETF", "sz.159636"),
    # ("纳指100ETF", "sh.513100"),
    # ("新能电池ETF", "sz.159755"),
    # ("电网设备ETF", "sz.159326"),
]
ETF_CONFIG = [
    ("中概互联ETF", "sz.159605"),
    ("中概互联网ETF", "sh.513050"),
    ("港股通科技30ETF", "sz.159636"),
    ("恒生科技ETF", "sh.513180"),
    ("港股通互联网ETF", "sz.159792"),
    # ("恒生互联网ETF", "sh.513330"),
    ("恒生港股通创新药ETF", "sz.159316"),
    ("中证港股通创新药ETF", "sh.513780"),
    ("中证创新药ETF", "sz.159992"),
    ("中韩半导体ETF", "sh.513310"),

    ("中证电网设备ETF", "sz.159326"),
    ("储能电池ETF", "sz.159566"),
    ("电池ETF", "sz.159755"),
    ("中证新能源ETF", "sh.516160"),
    ("半导体设备ETF", "sz.159516"),
    ("国证芯片ETF", "sz.159995"),
    ("半导体ETF", "sh.512480"),
    ("航空航天ETF", "sz.159227"),

    ("中证银行ETF", "sh.512800"),
    ("中证证券ETF", "sh.512880"),
    ("红利低波100ETF", "sh.515100"),
    ("中证红利低波ETF", "sh.512890"),

    ("黄金ETF", "sh.518880"),
    ("纳斯达克100ETF", "sh.513100"),
    ("标普500ETF", "sh.513650"),
    ("日经225ETF", "sh.513880"),

    ("消费电子ETF", "sz.159732"),
    ("TMT50ETF", "sz.159909"),
    ("机器人ETF", "sh.562500"),
    ("高端制造ETF", "sh.562910"),
    ("智能制造ETF", "sh.516800"),

    ("光伏ETF", "sh.515790"),
    ("通信设备ETF", "sh.515880"),
    ("5G通信ETF", "sh.515050"),
    ("消费50ETF", "sh.515650"),

    ("稀有金属ETF", "sh.562800"),
    ("有色金属ETF", "sh.512400"),
    ("化工ETF", "sz.159870"),
    ("有色ETF", "sz.159980"),
    ("能源化工ETF", "sz.159981"),

    ("深证100ETF", "sz.159901"),
    ("沪深300ETF", "sh.510300"),
    ("中证A500ETF", "sh.563360"),
    ("科创50ETF", "sh.588000"),
    ("创业板ETF", "sz.159915"),

    ("工业母机ETF", "sz.159667"),
    ("电力ETF", "sz.159611"),
    ("绿色电力ETF", "sz.159625"),
    ("家电ETF", "sz.159996"),
    ("港股通金融ETF", "sh.513190"),
    ("沪港深消费龙头ETF", "sh.517550"),
    ("人工智能50ETF", "sh.517800"),
    ("港股通高股息ETF", "sz.159691"),
    ("中证医疗ETF", "sh.512170"),
    ("CES芯片ETF", "sh.512760"),
    ("恒生消费ETF", "sh.513970"),
    ("中证新能源汽车ETF", "sh.515030"),
    ("食品饮料ETF", "sh.515170"),
    ("中证智能汽车ETF", "sh.516520"),
    # ("沪港深云计算ETF", "sz.159738"),
    # ("科创创业50ETF", "sz.159781"),
    # ("中证人工智能ETF", "sz.159819"),
    # ("互联网龙头ETF", "sz.159856"),
    # ("恒生ETF", "sz.159920"),
    # ("消费ETF", "sz.159928"),
    # ("金融地产ETF", "sz.159940"),
    # ("创业板50ETF", "sz.159949"),
    # ("创业板价值ETF", "sz.159966"),
    # ("创业板成长ETF", "sz.159967"),
    # ("生物医药ETF", "sh.512290"),
    # ("港股通50ETF", "sh.513550"),
    # ("标普港股低波红利ETF", "sh.513630"),
    # ("港股通非银ETF", "sh.513750"),
    # ("恒生A股电网设备ETF", "sh.561380"),
    # ("科创100ETF", "sh.588030"),
    # ("科创AIETF", "sh.588790"),
    # ("科创综指ETF", "sh.589000"),
    # ("恒生创新药ETF", "sh.520500"),
    # ("科技50ETF", "sh.515750"),
    # ("稀土ETF", "sh.516150"),
    # ("云计算大数据ETF", "sh.516510"),

    # ("国证自由现金流ETF", "sz.159201"),
    # ("全指红利质量ETF", "sz.159209"),
    # ("富时A股自由现金流ETF", "sz.159399"),
    # ("恒生高股息低波ETF", "sz.159545"),
    # ("中证红利质量ETF", "sz.159758"),
    # ("旅游ETF", "sz.159766"),
    # ("农业ETF", "sz.159825"),
    # ("软件ETF", "sz.159852"),
    # ("游戏ETF", "sz.159869"),
    # ("恒生医药ETF", "sz.159892"),
    # ("深证红利ETF", "sz.159905"),
    # ("电子ETF", "sz.159997"),
    # ("计算机ETF", "sz.159998"),
    # ("上证50ETF", "sh.510050"),
    # ("中证500ETF", "sh.510500"),
    # ("红利国企ETF", "sh.510720"),
    # ("恒生国企ETF", "sh.510900"),
    # ("医药ETF", "sh.512010"),
    # ("证券保险ETF", "sh.512070"),
    # ("中证1000ETF", "sh.512100"),
    # ("房地产ETF", "sh.512200"),
    # ("富时中国A50ETF", "sh.512550"),
    # ("中证军工ETF", "sh.512660"),
    # ("酒ETF", "sh.512690"),
    # ("军工龙头ETF", "sh.512710"),
    # ("央企改革ETF", "sh.512950"),
    # ("传媒ETF", "sh.512980"),
    # ("中证港股通央企红利ETF", "sh.513910"),
    # ("恒生港股通央企红利ETF", "sh.513920"),
    # ("科技ETF", "sh.515000"),
    # ("中证红利ETF", "sh.515180"),
    # ("钢铁ETF", "sh.515210"),
    # ("300红利低波ETF", "sh.515300"),
    # ("大数据ETF", "sh.515400"),
    # ("新能源汽车产业ETF", "sh.515700"),
    # ("消费服务ETF", "sh.516600"),
    # ("基建ETF", "sh.516970"),
    # ("央企共赢ETF", "sh.517090"),
    # ("中国国企ETF", "sh.517180"),
    # ("国新港股通央企红利ETF", "sh.520990"),
    # ("MSCI中国A50ETF", "sh.560050"),
    # ("央企科技ETF", "sh.560170"),
    # ("碳中和ETF", "sh.561190"),
    # ("中证2000ETF", "sh.563300"),
]

GLOBAL_START_DATE = "2026-01-05"
# END_DATE = "2026-07-11"
END_DATE = datetime.now().strftime("%Y-%m-%d")
#
# ========================== 网络超时与重试配置 ==========================
# 每一次 Baostock 完整处理最多等待 30 秒。超时后会强制结束子进程，
# 因此即使远程接口永久不返回，也不会卡住整个任务。
REQUEST_TIMEOUT_SECONDS = 30

# 第一轮中，每个标的最多尝试 3 次。
PRIMARY_MAX_ATTEMPTS = 3

# 第一轮全部结束后，对仍失败的标的统一再尝试 1 次。
FINAL_RETRY_ATTEMPTS = 1

# 两次尝试之间短暂等待，避免连续快速请求远程服务。
RETRY_INTERVAL_SECONDS = 2

# 预测标志：1=快金叉，-1=快死叉，0=预测下一期不会交叉
PREDICT_GOLDEN_CROSS = 1
PREDICT_DEATH_CROSS = -1
PREDICT_NO_CROSS = 0


# ========================== 基础工具 ==========================
def normalize_name(name: object) -> str:
    """名称比较时忽略空白并统一大小写。"""
    return re.sub(r"\s+", "", str(name or "")).upper()


def parse_config_item(item: Sequence[str]) -> Tuple[str, str]:
    """兼容二元组和三元组配置。"""
    if not isinstance(item, (tuple, list)) or len(item) < 2:
        raise ValueError(f"ETF_CONFIG 项格式错误：{item!r}，至少需要 (名称, 代码)")
    configured_name = str(item[0]).strip()
    code = str(item[1]).strip()
    if not configured_name or not code:
        raise ValueError(f"ETF_CONFIG 项存在空名称或空代码：{item!r}")
    return configured_name, code


def status_text(value: object) -> str:
    try:
        value = int(value)
    except (TypeError, ValueError):
        return "未知"
    if value == 1:
        return "金叉"
    if value == -1:
        return "死叉"
    return "中性"


def forecast_text(value: object) -> str:
    try:
        value = int(value)
    except (TypeError, ValueError):
        return "无法预测"
    if value == PREDICT_GOLDEN_CROSS:
        return "快金叉"
    if value == PREDICT_DEATH_CROSS:
        return "快死叉"
    return "无临近交叉"


def format_number(value: object, digits: int = 6) -> str:
    try:
        if pd.isna(value):
            return "NA"
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return "NA"


def make_unique_sheet_name(raw_name: str, used_names: set) -> str:
    """生成合法且唯一的 Excel Sheet 名称。"""
    clean = re.sub(r"[\\/*?:\[\]]", "_", str(raw_name)).strip()
    clean = clean or "Sheet"
    clean = clean[:31]

    candidate = clean
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{clean[:31 - len(suffix)]}{suffix}"
        index += 1

    used_names.add(candidate)
    return candidate


# ========================== 交叉预警与 Excel 颜色 ==========================
NEW_CROSS_MARKER = "★"

# 实际发生金叉/死叉：使用醒目的实色背景和白色粗体字。
ACTUAL_GOLDEN_FILL_COLOR = "00B050"   # 绿色
ACTUAL_DEATH_FILL_COLOR = "C00000"    # 红色
ACTUAL_FONT_COLOR = "FFFFFF"

# 临近金叉/死叉：使用淡色背景，避免与已经发生的交叉混淆。
NEAR_GOLDEN_FILL_COLOR = "E2F0D9"     # 淡绿色
NEAR_GOLDEN_FONT_COLOR = "006100"
NEAR_DEATH_FILL_COLOR = "FCE4D6"      # 淡红色
NEAR_DEATH_FONT_COLOR = "9C0006"


def is_first_cross_period(value: object) -> bool:
    """持续时间是否为 1；兼容 int、float、字符串和空值。"""
    try:
        if pd.isna(value):
            return False
        return int(float(value)) == 1
    except (TypeError, ValueError):
        return False


def add_console_cross_marker(status: object, age: object) -> str:
    """持续时间为 1 时，在控制台状态文字前加 ★。"""
    text = str(status)
    return f"{NEW_CROSS_MARKER}{text}" if is_first_cross_period(age) else text


def _safe_int(value: object) -> Optional[int]:
    """将 Excel/Pandas 数值安全转换为 int，转换失败时返回 None。"""
    try:
        if pd.isna(value):
            return None
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _apply_actual_cross_style(cell, signal: int) -> None:
    """实际金叉用绿色，实际死叉用红色。"""
    if signal == 1:
        fill_color = ACTUAL_GOLDEN_FILL_COLOR
    elif signal == -1:
        fill_color = ACTUAL_DEATH_FILL_COLOR
    else:
        return

    cell.font = Font(color=ACTUAL_FONT_COLOR, bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)


def _apply_near_cross_style(cell, signal: int) -> None:
    """临近金叉用淡绿色，临近死叉用淡红色。"""
    if signal == 1:
        font_color = NEAR_GOLDEN_FONT_COLOR
        fill_color = NEAR_GOLDEN_FILL_COLOR
    elif signal == -1:
        font_color = NEAR_DEATH_FONT_COLOR
        fill_color = NEAR_DEATH_FILL_COLOR
    else:
        return

    cell.font = Font(color=font_color, bold=True)
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)


def apply_summary_cross_color_format(
    writer: pd.ExcelWriter,
    summary_df: pd.DataFrame,
    sheet_name: str = "汇总",
) -> None:
    """
    设置汇总 Sheet 的颜色：

    1. 当日/本周刚金叉：当前状态、交叉事件、持续时间、持续说明标绿色。
    2. 当日/本周刚死叉：上述单元格标红色。
    3. 快金叉：预测文字和预测标志标淡绿色。
    4. 快死叉：预测文字和预测标志标淡红色。
    """
    worksheet = writer.sheets[sheet_name]
    header_to_column = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    indicator_groups = [
        {
            "status": "日MACD当前",
            "event": "日MACD交叉事件",
            "duration": "日MACD已持续交易日",
            "age_text": "日MACD持续说明",
            "forecast": "日MACD预测",
            "forecast_flag": "日MACD预测标志",
        },
        {
            "status": "日KDJ当前",
            "event": "日KDJ交叉事件",
            "duration": "日KDJ已持续交易日",
            "age_text": "日KDJ持续说明",
            "forecast": "日KDJ预测",
            "forecast_flag": "日KDJ预测标志",
        },
        {
            "status": "周MACD当前",
            "event": "周MACD交叉事件",
            "duration": "周MACD已持续周数",
            "age_text": "周MACD持续说明",
            "forecast": "周MACD预测",
            "forecast_flag": "周MACD预测标志",
        },
        {
            "status": "周KDJ当前",
            "event": "周KDJ交叉事件",
            "duration": "周KDJ已持续周数",
            "age_text": "周KDJ持续说明",
            "forecast": "周KDJ预测",
            "forecast_flag": "周KDJ预测标志",
        },
    ]

    for dataframe_row_index, row in summary_df.iterrows():
        excel_row = int(dataframe_row_index) + 2

        for group in indicator_groups:
            # 实际新交叉：持续时间为 1，方向由“当前”字段判断。
            if group["duration"] in summary_df.columns and is_first_cross_period(
                row[group["duration"]]
            ):
                status = str(row.get(group["status"], ""))
                actual_signal = 1 if status == "金叉" else -1 if status == "死叉" else 0

                for column_name in (
                    group["status"],
                    group["event"],
                    group["duration"],
                    group["age_text"],
                ):
                    excel_column = header_to_column.get(column_name)
                    if excel_column is not None:
                        _apply_actual_cross_style(
                            worksheet.cell(row=excel_row, column=excel_column),
                            actual_signal,
                        )

            # 临近交叉：直接使用预测标志 1/-1。
            forecast_signal = _safe_int(row.get(group["forecast_flag"]))
            if forecast_signal in (1, -1):
                for column_name in (group["forecast"], group["forecast_flag"]):
                    excel_column = header_to_column.get(column_name)
                    if excel_column is not None:
                        _apply_near_cross_style(
                            worksheet.cell(row=excel_row, column=excel_column),
                            forecast_signal,
                        )


def apply_detail_cross_color_format(
    writer: pd.ExcelWriter,
    sheet_name: str,
) -> None:
    """
    设置每个 ETF 明细 Sheet 的颜色。

    实际交叉依据 *_cross_event：
      1  -> 绿色
      -1 -> 红色

    临近交叉依据 *_near_cross_flag：
      1  -> 淡绿色
      -1 -> 淡红色
    """
    worksheet = writer.sheets[sheet_name]
    header_to_column = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    groups = [
        {
            "prefix": "day_macd",
            "age": "day_macd_cross_days",
        },
        {
            "prefix": "day_kdj",
            "age": "day_kdj_cross_days",
        },
        {
            "prefix": "week_macd",
            "age": "week_macd_cross_weeks",
        },
        {
            "prefix": "week_kdj",
            "age": "week_kdj_cross_weeks",
        },
    ]

    for group in groups:
        prefix = group["prefix"]
        actual_event_col = f"{prefix}_cross_event"
        actual_columns = [
            f"{prefix}_status",
            actual_event_col,
            f"{prefix}_cross_event_text",
            group["age"],
            f"{prefix}_cross_age_text",
        ]
        forecast_flag_col = f"{prefix}_near_cross_flag"
        forecast_columns = [
            forecast_flag_col,
            f"{prefix}_near_cross_text",
        ]

        event_excel_col = header_to_column.get(actual_event_col)
        forecast_excel_col = header_to_column.get(forecast_flag_col)

        for excel_row in range(2, worksheet.max_row + 1):
            if event_excel_col is not None:
                actual_signal = _safe_int(
                    worksheet.cell(row=excel_row, column=event_excel_col).value
                )
                if actual_signal in (1, -1):
                    for column_name in actual_columns:
                        excel_column = header_to_column.get(column_name)
                        if excel_column is not None:
                            _apply_actual_cross_style(
                                worksheet.cell(row=excel_row, column=excel_column),
                                actual_signal,
                            )

            if forecast_excel_col is not None:
                forecast_signal = _safe_int(
                    worksheet.cell(row=excel_row, column=forecast_excel_col).value
                )
                if forecast_signal in (1, -1):
                    for column_name in forecast_columns:
                        excel_column = header_to_column.get(column_name)
                        if excel_column is not None:
                            _apply_near_cross_style(
                                worksheet.cell(row=excel_row, column=excel_column),
                                forecast_signal,
                            )


# ========================== 远程证券信息 ==========================
def query_remote_basic_info(code: str) -> Dict[str, Optional[str]]:
    """
    从 Baostock 远程查询证券名称和上市日期。

    返回字段：
      remote_code, remote_name, ipo_date, error
    """
    result: Dict[str, Optional[str]] = {
        "remote_code": code,
        "remote_name": None,
        "ipo_date": None,
        "error": None,
    }

    rs = bs.query_stock_basic(code=code)
    if rs.error_code != "0":
        result["error"] = f"{rs.error_code}: {rs.error_msg}"
        return result

    if not rs.next():
        result["error"] = "远程未返回证券基本信息"
        return result

    fields = list(getattr(rs, "fields", []) or [])
    row_data = rs.get_row_data()
    row = dict(zip(fields, row_data))

    # Baostock 官方字段通常为 code、code_name、ipoDate 等。
    remote_code = (row.get("code") or code).strip()
    remote_name = (row.get("code_name") or row.get("name") or "").strip()
    ipo_raw = (row.get("ipoDate") or "").strip()

    result["remote_code"] = remote_code or code
    result["remote_name"] = remote_name or None

    if ipo_raw:
        ipo_date = pd.to_datetime(ipo_raw, errors="coerce")
        if not pd.isna(ipo_date):
            result["ipo_date"] = ipo_date.strftime("%Y-%m-%d")

    return result


# ========================== 技术指标 ==========================
def calculate_macd(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    ema_fast = df["close"].ewm(span=10, adjust=False).mean()
    ema_slow = df["close"].ewm(span=22, adjust=False).mean()
    dif = ema_fast - ema_slow
    dea = dif.ewm(span=9, adjust=False).mean()
    macd = (dif - dea) * 2

    df[f"{prefix}dif"] = dif
    df[f"{prefix}dea"] = dea
    df[f"{prefix}macd"] = macd
    return df


def add_macd_zero_axis_status(
    df: pd.DataFrame,
    dif_col: str,
    dea_col: str,
    output_col: str,
) -> pd.DataFrame:
    """
    根据 MACD 的 DIF、DEA 与零轴的位置，生成水上/水下状态。

    判定规则：
      DIF >= 0 且 DEA >= 0  -> 完全水上
      DIF <= 0 且 DEA <= 0  -> 完全水下
      DIF < 0  且 DEA > 0   -> 水上到水下过渡
      DIF > 0  且 DEA < 0   -> 水下到水上过渡

    DIF 和 DEA 同时等于 0 时归入“完全水上”。该情况通常只出现在
    MACD 初始化数据的最前端，不影响最新周线状态判断。
    """
    dif = pd.to_numeric(df[dif_col], errors="coerce")
    dea = pd.to_numeric(df[dea_col], errors="coerce")
    valid = dif.notna() & dea.notna()

    status = pd.Series("未知", index=df.index, dtype="object")

    completely_above = valid & (dif >= 0) & (dea >= 0)
    completely_below = (
        valid
        & (dif <= 0)
        & (dea <= 0)
        & ~completely_above
    )
    above_to_below = valid & (dif < 0) & (dea > 0)
    below_to_above = valid & (dif > 0) & (dea < 0)

    status.loc[completely_above] = "完全水上"
    status.loc[completely_below] = "完全水下"
    status.loc[above_to_below] = "水上到水下过渡"
    status.loc[below_to_above] = "水下到水上过渡"

    df[output_col] = status
    return df


def calculate_kdj(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    low_list = df["low"].rolling(9, min_periods=1).min()
    high_list = df["high"].rolling(9, min_periods=1).max()

    hl_range = (high_list - low_list).mask((high_list - low_list) == 0)
    rsv = (df["close"] - low_list) / hl_range * 100
    rsv = rsv.fillna(50).ffill()

    k = rsv.ewm(com=2, adjust=False).mean()
    d = k.ewm(com=2, adjust=False).mean()
    j = 3 * k - 2 * d

    df[f"{prefix}k"] = k
    df[f"{prefix}d"] = d
    df[f"{prefix}j"] = j
    return df


def add_cross_status_metrics(
    df: pd.DataFrame,
    left_col: str,
    right_col: str,
    prefix: str,
    age_col: str,
    period_label: str,
) -> pd.DataFrame:
    """
    计算金叉/死叉当前状态、交叉事件和当前状态持续期数。

    生成字段：
      {prefix}_status
        1  = 左线在右线上方，当前处于金叉状态
        -1 = 左线在右线下方，当前处于死叉状态
        0  = 两线始终相等，无法判断

      {prefix}_cross_event
        1  = 本期刚刚由死叉转为金叉
        -1 = 本期刚刚由金叉转为死叉
        0  = 本期没有发生新的交叉

      {age_col}
        当前金叉/死叉状态已经连续保持的期数。
        交叉发生当期为 1，下一期若状态不变则为 2，以此类推。

      {prefix}_cross_age_text
        例如“金叉第3个交易日”“死叉第2周”。

    注意：如果数据起点之前已经发生交叉，数据第一段状态的持续期数只能
    从当前数据集起点开始计算；后续发生的交叉可以准确从 1 开始计数。
    """
    status_col = f"{prefix}_status"
    event_col = f"{prefix}_cross_event"
    event_text_col = f"{prefix}_cross_event_text"
    age_text_col = f"{prefix}_cross_age_text"

    gap = df[left_col] - df[right_col]

    status = pd.Series(pd.NA, index=df.index, dtype="Int64")
    status.loc[gap > 0] = 1
    status.loc[gap < 0] = -1

    # 两线暂时相等时沿用最近状态；开头连续相等时使用后续第一个有效状态。
    status = status.ffill().bfill().fillna(0).astype(int)
    df[status_col] = status

    previous_status = status.shift(1)
    event = pd.Series(0, index=df.index, dtype="int64")
    event.loc[(status == 1) & (previous_status == -1)] = 1
    event.loc[(status == -1) & (previous_status == 1)] = -1
    df[event_col] = event

    if period_label == "周":
        event_text_map = {1: "本周金叉", -1: "本周死叉", 0: "本周无新交叉"}
    else:
        event_text_map = {1: "当日金叉", -1: "当日死叉", 0: "当日无新交叉"}
    df[event_text_col] = df[event_col].map(event_text_map)

    # 每当状态变化就开启一个新分组；组内序号从 1 开始。
    run_id = status.ne(status.shift(1)).cumsum()
    age = status.groupby(run_id).cumcount() + 1
    age = age.where(status != 0, 0).astype(int)
    df[age_col] = age

    age_text = []
    for current_status, current_age in zip(status.tolist(), age.tolist()):
        if current_status == 1:
            age_text.append(f"金叉第{current_age}{period_label}")
        elif current_status == -1:
            age_text.append(f"死叉第{current_age}{period_label}")
        else:
            age_text.append("中性")
    df[age_text_col] = age_text

    return df


def add_linear_cross_forecast(
    df: pd.DataFrame,
    left_col: str,
    right_col: str,
    prefix: str,
) -> pd.DataFrame:
    """
    使用最近两期“线差”做一次线性外推。

    gap_t      = left_t - right_t
    next_gap   = gap_t + (gap_t - gap_t-1)
               = 2 * gap_t - gap_t-1

    判断：
      当前 gap <= 0 且预测 next_gap > 0  -> 快金叉，flag=1
      当前 gap >= 0 且预测 next_gap < 0  -> 快死叉，flag=-1
      否则                              -> 无临近交叉，flag=0
    """
    gap_col = f"{prefix}_gap"
    previous_gap_col = f"{prefix}_prev_gap"
    predicted_gap_col = f"{prefix}_next_gap_pred"
    flag_col = f"{prefix}_near_cross_flag"
    text_col = f"{prefix}_near_cross_text"

    df[gap_col] = df[left_col] - df[right_col]
    df[previous_gap_col] = df[gap_col].shift(1)
    df[predicted_gap_col] = 2 * df[gap_col] - df[previous_gap_col]

    flag = pd.Series(PREDICT_NO_CROSS, index=df.index, dtype="int64")
    valid = df[previous_gap_col].notna() & df[predicted_gap_col].notna()

    golden = valid & (df[gap_col] <= 0) & (df[predicted_gap_col] > 0)
    death = valid & (df[gap_col] >= 0) & (df[predicted_gap_col] < 0)

    flag.loc[golden] = PREDICT_GOLDEN_CROSS
    flag.loc[death] = PREDICT_DEATH_CROSS

    df[flag_col] = flag
    df[text_col] = df[flag_col].map(
        {
            PREDICT_GOLDEN_CROSS: "快金叉",
            PREDICT_DEATH_CROSS: "快死叉",
            PREDICT_NO_CROSS: "无临近交叉",
        }
    )
    return df


# ========================== 周线生成与预测 ==========================
def add_weekly_indicators(daily_df: pd.DataFrame) -> pd.DataFrame:
    df = daily_df.copy()

    if not pd.api.types.is_datetime64_any_dtype(df["date"]):
        print("  检测到 date 不是 datetime，正在转换...")
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).reset_index(drop=True)
        if df.empty:
            raise ValueError("所有日期转换失败，无法生成周线")

    df["date"] = df["date"].dt.normalize()

    weekly_df = (
        df.resample("W-FRI", on="date")
        .agg(
            {
                "open": "first",
                "high": "max",
                "low": "min",
                "close": "last",
                "volume": "sum",
                "amount": "sum",
            }
        )
        .reset_index()
    )

    # 去掉整周无交易造成的空周，避免技术指标被空值污染。
    weekly_df = weekly_df.dropna(subset=["close"]).reset_index(drop=True)
    weekly_df["date"] = weekly_df["date"].dt.normalize()

    weekly_df = calculate_macd(weekly_df, prefix="week_")

    # 周 MACD 水上/水下状态：根据周 DIF、DEA 相对零轴的位置判断。
    weekly_df = add_macd_zero_axis_status(
        weekly_df,
        dif_col="week_dif",
        dea_col="week_dea",
        output_col="week_macd_water_status",
    )

    weekly_df = calculate_kdj(weekly_df, prefix="week_")

    # 周线状态、交叉事件和持续周数必须在“唯一周线数据”上计算，
    # 不能在重复映射后的日线上计算，否则持续周数会被交易日重复放大。
    weekly_df = add_cross_status_metrics(
        weekly_df,
        left_col="week_dif",
        right_col="week_dea",
        prefix="week_macd",
        age_col="week_macd_cross_weeks",
        period_label="周",
    )
    weekly_df = add_cross_status_metrics(
        weekly_df,
        left_col="week_k",
        right_col="week_d",
        prefix="week_kdj",
        age_col="week_kdj_cross_weeks",
        period_label="周",
    )

    # 周 MACD 和周 KDJ：使用最近两个周线差值预测下一周。
    weekly_df = add_linear_cross_forecast(
        weekly_df, "week_dif", "week_dea", "week_macd"
    )
    weekly_df = add_linear_cross_forecast(
        weekly_df, "week_k", "week_d", "week_kdj"
    )

    df["week_end_date"] = (
        df["date"].dt.to_period("W-FRI").dt.end_time.dt.normalize()
    )

    weekly_merge = weekly_df.rename(columns={"date": "week_end_date"}).copy()
    weekly_columns = [
        "week_end_date",
        "week_macd_water_status",
        "week_dif",
        "week_dea",
        "week_macd",
        "week_k",
        "week_d",
        "week_j",
        "week_macd_status",
        "week_macd_cross_event",
        "week_macd_cross_event_text",
        "week_macd_cross_weeks",
        "week_macd_cross_age_text",
        "week_kdj_status",
        "week_kdj_cross_event",
        "week_kdj_cross_event_text",
        "week_kdj_cross_weeks",
        "week_kdj_cross_age_text",
        "week_macd_gap",
        "week_macd_prev_gap",
        "week_macd_next_gap_pred",
        "week_macd_near_cross_flag",
        "week_macd_near_cross_text",
        "week_kdj_gap",
        "week_kdj_prev_gap",
        "week_kdj_next_gap_pred",
        "week_kdj_near_cross_flag",
        "week_kdj_near_cross_text",
    ]

    df = pd.merge(
        df,
        weekly_merge[weekly_columns],
        on="week_end_date",
        how="left",
        validate="many_to_one",
    )
    df = df.drop(columns=["week_end_date"])

    matched_count = int(df["week_dif"].notna().sum())
    print(f"  周线指标匹配成功：{matched_count}/{len(df)} 行")
    return df


# ========================== 单标的数据获取 ==========================
def get_etf_full_data(code: str, configured_name: str) -> Optional[pd.DataFrame]:
    print(f"\n{'=' * 72}")
    print(f"配置名称：{configured_name}")
    print(f"证券代码：{code}")
    print("-" * 72)

    login_result = bs.login()
    if login_result.error_code != "0":
        print(f"Baostock 登录失败：{login_result.error_msg}")
        return None

    try:
        basic = query_remote_basic_info(code)
        remote_name = basic.get("remote_name")
        remote_code = basic.get("remote_code") or code
        ipo_date = basic.get("ipo_date")

        if remote_name:
            name_match = normalize_name(configured_name) == normalize_name(remote_name)
            display_name = remote_name
            print(f"远程名称：{remote_name}")
            if name_match:
                print("名称校验：一致")
            else:
                print(
                    f"名称校验：不一致。配置名称={configured_name}，"
                    f"远程名称={remote_name}；后续自动使用远程名称。"
                )
        else:
            name_match = False
            display_name = configured_name
            print(
                f"远程名称：获取失败（{basic.get('error') or '未知原因'}）；"
                "后续暂用配置名称。"
            )

        if remote_code != code:
            print(f"代码校验：远程返回代码 {remote_code}，与配置代码 {code} 不一致")
        else:
            print("代码校验：一致")

        start_date = GLOBAL_START_DATE
        if ipo_date:
            start_date = max(GLOBAL_START_DATE, ipo_date)
            print(f"上市日期：{ipo_date}，实际起始日期：{start_date}")
        else:
            print(f"上市日期：未获取，使用全局起始日期 {GLOBAL_START_DATE}")

        print(f"请求数据时间范围：{start_date} 至 {END_DATE}")

        rs = bs.query_history_k_data_plus(
            code=code,
            fields="date,open,high,low,close,volume,amount,pctChg,turn",
            start_date=start_date,
            end_date=END_DATE,
            frequency="d",
            adjustflag="2",
        )

        if rs.error_code != "0":
            print(f"获取行情数据失败：{rs.error_msg}")
            return None

        data_list: List[List[str]] = []
        while rs.next():
            data_list.append(rs.get_row_data())

        if not data_list:
            print("未获取到任何行情数据")
            return None

        df = pd.DataFrame(data_list, columns=rs.fields)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).reset_index(drop=True)

        if df.empty:
            print("所有日期转换失败")
            return None

        numeric_columns = [col for col in df.columns if col != "date"]
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        # 核心行情列缺失的行不能参与指标计算。
        df = df.dropna(subset=["open", "high", "low", "close"]).reset_index(drop=True)
        if df.empty:
            print("有效 OHLC 数据为空")
            return None

        print(f"原始数据获取成功：共 {len(df)} 个交易日")
        print(f"数据起始日期：{df['date'].iloc[0].strftime('%Y-%m-%d')}")
        print(f"数据结束日期：{df['date'].iloc[-1].strftime('%Y-%m-%d')}")

        print("正在计算日线 MACD/KDJ...")
        df = calculate_macd(df, prefix="day_")

        # 日 MACD 水上/水下状态：根据日 DIF、DEA 相对零轴的位置判断。
        df = add_macd_zero_axis_status(
            df,
            dif_col="day_dif",
            dea_col="day_dea",
            output_col="day_macd_water_status",
        )

        df = calculate_kdj(df, prefix="day_")

        df = add_cross_status_metrics(
            df,
            left_col="day_dif",
            right_col="day_dea",
            prefix="day_macd",
            age_col="day_macd_cross_days",
            period_label="个交易日",
        )
        df = add_cross_status_metrics(
            df,
            left_col="day_k",
            right_col="day_d",
            prefix="day_kdj",
            age_col="day_kdj_cross_days",
            period_label="个交易日",
        )

        # 日 MACD、日 KDJ：使用最近两个交易日的线差预测下一交易日。
        # 例如上期差=1、本期差=3，则预测下期差=3+(3-1)=5。
        df = add_linear_cross_forecast(df, "day_dif", "day_dea", "day_macd")
        df = add_linear_cross_forecast(df, "day_k", "day_d", "day_kdj")

        print("正在计算周线 MACD/KDJ，并预测下一周交叉...")
        df = add_weekly_indicators(df)

        # 把远程名称校验信息写入每一行，便于 Excel 中直接核对。
        df.insert(0, "name_match", bool(remote_name) and name_match)
        df.insert(0, "remote_name", remote_name or "")
        df.insert(0, "configured_name", configured_name)
        df.insert(0, "code", code)

        df.attrs["code"] = code
        df.attrs["configured_name"] = configured_name
        df.attrs["remote_name"] = remote_name or ""
        df.attrs["display_name"] = display_name
        df.attrs["name_match"] = bool(remote_name) and name_match
        df.attrs["name_verified"] = bool(remote_name)

        latest = df.iloc[-1]
        rate = (latest["close"] / df["close"].iloc[0] - 1) * 100

        print(f"\n{display_name}（{code}）处理完成")
        print(f"总数据条数：{len(df)}")
        print(f"累计涨跌幅：{rate:.2f}%")
        print(f"最新日期：{latest['date'].strftime('%Y-%m-%d')}")
        print(
            f"日线 MACD 水上水下状态：{latest['day_macd_water_status']}"
        )
        print(
            f"日线 MACD 当前：{status_text(latest['day_macd_status'])}，"
            f"已持续 {int(latest['day_macd_cross_days'])} 个交易日"
        )
        print(
            f"日线 KDJ 当前：{status_text(latest['day_kdj_status'])}，"
            f"已持续 {int(latest['day_kdj_cross_days'])} 个交易日"
        )
        print(
            f"周线 MACD 水上水下状态：{latest['week_macd_water_status']}"
        )
        print(
            f"周线 MACD 当前：{status_text(latest['week_macd_status'])}，"
            f"已持续 {int(latest['week_macd_cross_weeks'])} 周"
        )
        print(
            f"周线 KDJ 当前：{status_text(latest['week_kdj_status'])}，"
            f"已持续 {int(latest['week_kdj_cross_weeks'])} 周"
        )

        print("\n线性外推结果：")
        print(
            "  日 MACD（预测下一交易日）："
            f"上期差={format_number(latest['day_macd_prev_gap'])}，"
            f"本期差={format_number(latest['day_macd_gap'])}，"
            f"预测下期差={format_number(latest['day_macd_next_gap_pred'])}，"
            f"标志={latest['day_macd_near_cross_text']}"
        )
        print(
            "  日 KDJ（预测下一交易日）："
            f"上期差={format_number(latest['day_kdj_prev_gap'])}，"
            f"本期差={format_number(latest['day_kdj_gap'])}，"
            f"预测下期差={format_number(latest['day_kdj_next_gap_pred'])}，"
            f"标志={latest['day_kdj_near_cross_text']}"
        )
        print(
            "  周 KDJ（预测下一周）："
            f"上周差={format_number(latest['week_kdj_prev_gap'])}，"
            f"本周差={format_number(latest['week_kdj_gap'])}，"
            f"预测下周差={format_number(latest['week_kdj_next_gap_pred'])}，"
            f"标志={latest['week_kdj_near_cross_text']}"
        )
        print(
            "  周 MACD（预测下一周）："
            f"上周差={format_number(latest['week_macd_prev_gap'])}，"
            f"本周差={format_number(latest['week_macd_gap'])}，"
            f"预测下周差={format_number(latest['week_macd_next_gap_pred'])}，"
            f"标志={latest['week_macd_near_cross_text']}"
        )

        return df

    except Exception as exc:
        print(f"处理失败：{exc}")
        traceback.print_exc()
        return None

    finally:
        try:
            bs.logout()
        except Exception:
            pass


# ========================== 汇总与导出 ==========================
def build_summary_row(df: pd.DataFrame) -> Dict[str, object]:
    latest = df.iloc[-1]
    remote_name = df.attrs.get("remote_name", "")
    configured_name = df.attrs.get("configured_name", "")
    display_name = df.attrs.get("display_name", remote_name or configured_name)
    name_verified = bool(df.attrs.get("name_verified", False))
    name_match = bool(df.attrs.get("name_match", False))

    if not name_verified:
        name_check = "远程名称未获取"
    elif name_match:
        name_check = "一致"
    else:
        name_check = "不一致，已使用远程名称"

    return {
        "远程标的名称": display_name,
        "配置名称": configured_name,
        "证券代码": df.attrs.get("code", ""),
        "名称校验": name_check,
        "起始日期": df["date"].iloc[0].strftime("%Y-%m-%d"),
        "最新日期": latest["date"].strftime("%Y-%m-%d"),
        "数据条数": len(df),
        "日MACD水上水下状态": latest["day_macd_water_status"],
        "日MACD当前": status_text(latest["day_macd_status"]),
        "日MACD交叉事件": latest["day_macd_cross_event_text"],
        "日MACD已持续交易日": int(latest["day_macd_cross_days"]),
        "日MACD持续说明": latest["day_macd_cross_age_text"],
        "日MACD预测": latest["day_macd_near_cross_text"],
        "日MACD预测标志": int(latest["day_macd_near_cross_flag"]),
        "日MACD上期差": latest["day_macd_prev_gap"],
        "日MACD本期差": latest["day_macd_gap"],
        "日MACD预测下期差": latest["day_macd_next_gap_pred"],
        "日KDJ当前": status_text(latest["day_kdj_status"]),
        "日KDJ交叉事件": latest["day_kdj_cross_event_text"],
        "日KDJ已持续交易日": int(latest["day_kdj_cross_days"]),
        "日KDJ持续说明": latest["day_kdj_cross_age_text"],
        "日KDJ预测": latest["day_kdj_near_cross_text"],
        "日KDJ预测标志": int(latest["day_kdj_near_cross_flag"]),
        "日KDJ上期差": latest["day_kdj_prev_gap"],
        "日KDJ本期差": latest["day_kdj_gap"],
        "日KDJ预测下期差": latest["day_kdj_next_gap_pred"],
        "周KDJ当前": status_text(latest["week_kdj_status"]),
        "周KDJ交叉事件": latest["week_kdj_cross_event_text"],
        "周KDJ已持续周数": int(latest["week_kdj_cross_weeks"]),
        "周KDJ持续说明": latest["week_kdj_cross_age_text"],
        "周KDJ预测": latest["week_kdj_near_cross_text"],
        "周KDJ预测标志": int(latest["week_kdj_near_cross_flag"]),
        "周KDJ上周差": latest["week_kdj_prev_gap"],
        "周KDJ本周差": latest["week_kdj_gap"],
        "周KDJ预测下周差": latest["week_kdj_next_gap_pred"],
        "周MACD水上水下状态": latest["week_macd_water_status"],
        "周MACD当前": status_text(latest["week_macd_status"]),
        "周MACD交叉事件": latest["week_macd_cross_event_text"],
        "周MACD已持续周数": int(latest["week_macd_cross_weeks"]),
        "周MACD持续说明": latest["week_macd_cross_age_text"],
        "周MACD预测": latest["week_macd_near_cross_text"],
        "周MACD预测标志": int(latest["week_macd_near_cross_flag"]),
        "周MACD上周差": latest["week_macd_prev_gap"],
        "周MACD本周差": latest["week_macd_gap"],
        "周MACD预测下周差": latest["week_macd_next_gap_pred"],
    }


def format_elapsed_time(seconds: float) -> str:
    """
    将秒数格式化为中文时间。

    示例：
      30        -> 30秒
      80        -> 1分20秒
      3665      -> 1小时1分5秒
    """
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    if hours > 0:
        return f"{hours}小时{minutes}分{seconds}秒"
    if minutes > 0:
        return f"{minutes}分{seconds}秒"
    return f"{seconds}秒"



def _etf_process_worker(
    send_connection,
    code: str,
    configured_name: str,
) -> None:
    """
    子进程工作函数。

    每次尝试都在独立进程中完成登录、查询、指标计算和登出。
    父进程若等待超过 REQUEST_TIMEOUT_SECONDS，会直接结束该子进程，
    从而避免 Baostock 网络调用永久阻塞整个任务。
    """
    try:
        df = get_etf_full_data(code, configured_name)

        if df is None or df.empty:
            send_connection.send(
                (
                    "failed",
                    None,
                    "未获取到有效数据；可能是登录、证券信息或历史行情查询失败",
                )
            )
        else:
            send_connection.send(("success", df, ""))

    except BaseException as exc:
        error_message = (
            f"{type(exc).__name__}: {exc}\n"
            f"{traceback.format_exc()}"
        )
        try:
            send_connection.send(("error", None, error_message))
        except Exception:
            pass

    finally:
        try:
            send_connection.close()
        except Exception:
            pass


def _force_stop_process(process: mp.Process) -> None:
    """可靠结束超时或异常的子进程，避免残留进程继续占用连接。"""
    if not process.is_alive():
        process.join(timeout=1)
        return

    process.terminate()
    process.join(timeout=3)

    # Python 3.7+ 的 multiprocessing.Process 通常提供 kill()。
    if process.is_alive() and hasattr(process, "kill"):
        process.kill()
        process.join(timeout=3)


def run_etf_once_with_timeout(
    code: str,
    configured_name: str,
    timeout_seconds: int = REQUEST_TIMEOUT_SECONDS,
) -> Tuple[Optional[pd.DataFrame], str, str]:
    """
    对一个 ETF 执行一次有硬超时保护的处理。

    返回：
      dataframe, status, reason

    status：
      success = 成功
      timeout = 超过 timeout_seconds 未返回
      failed  = 函数正常返回，但没有有效数据
      error   = 子进程抛出异常
      crashed = 子进程异常退出且没有返回结果
    """
    # 显式使用 spawn，兼容 Windows / PyCharm，也避免继承异常网络连接状态。
    context = mp.get_context("spawn")
    receive_connection, send_connection = context.Pipe(duplex=False)

    process = context.Process(
        target=_etf_process_worker,
        args=(send_connection, code, configured_name),
        name=f"Baostock-{code}",
    )

    try:
        process.start()
        send_connection.close()

        deadline = time.monotonic() + timeout_seconds

        while True:
            remaining = deadline - time.monotonic()

            if remaining <= 0:
                _force_stop_process(process)
                return (
                    None,
                    "timeout",
                    f"超过 {timeout_seconds} 秒仍未返回，已强制结束本次查询",
                )

            # 分段等待，以便及时发现子进程已异常退出。
            if receive_connection.poll(min(0.2, remaining)):
                try:
                    status, df, reason = receive_connection.recv()
                except EOFError:
                    _force_stop_process(process)
                    return (
                        None,
                        "crashed",
                        "子进程连接提前关闭，没有返回结果",
                    )

                process.join(timeout=2)
                if process.is_alive():
                    _force_stop_process(process)

                return df, str(status), str(reason)

            if not process.is_alive():
                # 子进程可能已经发送数据，最后再检查一次管道。
                if receive_connection.poll(0):
                    try:
                        status, df, reason = receive_connection.recv()
                        process.join(timeout=1)
                        return df, str(status), str(reason)
                    except EOFError:
                        pass

                exit_code = process.exitcode
                process.join(timeout=1)
                return (
                    None,
                    "crashed",
                    f"子进程异常退出，exitcode={exit_code}，没有返回结果",
                )

    except BaseException as exc:
        _force_stop_process(process)
        return (
            None,
            "error",
            f"启动或管理子进程失败：{type(exc).__name__}: {exc}",
        )

    finally:
        try:
            receive_connection.close()
        except Exception:
            pass
        try:
            send_connection.close()
        except Exception:
            pass


def run_etf_with_retries(
    code: str,
    configured_name: str,
    max_attempts: int,
    phase_name: str,
    progress_bar=None,
) -> Tuple[Optional[pd.DataFrame], str, int]:
    """
    在指定阶段内重试一个 ETF。

    每次尝试都有独立的 30 秒硬超时；前一次超时后，子进程会被结束，
    不会留在后台继续阻塞。
    """
    last_reason = "未知原因"

    for attempt in range(1, max_attempts + 1):
        attempt_start = time.perf_counter()

        if progress_bar is not None:
            progress_bar.set_description_str(
                f"{phase_name}：{configured_name}（{code}）"
            )
            progress_bar.set_postfix_str(
                f"第 {attempt}/{max_attempts} 次｜单次上限 {REQUEST_TIMEOUT_SECONDS}秒"
            )
            progress_bar.refresh()

        tqdm.write(
            f"\n↻ {phase_name}｜{configured_name}（{code}）｜"
            f"开始第 {attempt}/{max_attempts} 次尝试，"
            f"最长等待 {REQUEST_TIMEOUT_SECONDS} 秒"
        )

        df, status, reason = run_etf_once_with_timeout(
            code=code,
            configured_name=configured_name,
            timeout_seconds=REQUEST_TIMEOUT_SECONDS,
        )

        attempt_elapsed = time.perf_counter() - attempt_start

        if status == "success" and df is not None and not df.empty:
            tqdm.write(
                f"✓ {configured_name}（{code}）第 {attempt} 次尝试成功｜"
                f"耗时 {format_elapsed_time(attempt_elapsed)}"
            )
            return df, "", attempt

        last_reason = f"{status}: {reason}"
        tqdm.write(
            f"⚠ {configured_name}（{code}）第 {attempt}/{max_attempts} 次失败｜"
            f"耗时 {format_elapsed_time(attempt_elapsed)}｜{last_reason}"
        )

        if attempt < max_attempts:
            tqdm.write(
                f"  等待 {RETRY_INTERVAL_SECONDS} 秒后重新连接并尝试下一次..."
            )
            time.sleep(RETRY_INTERVAL_SECONDS)

    return None, last_reason, max_attempts


def build_failure_dataframe(
    final_failures: List[Dict[str, object]],
    config_failures: List[Dict[str, object]],
) -> pd.DataFrame:
    """生成最终失败清单，用于控制台提示和 Excel 的“失败列表”Sheet。"""
    records: List[Dict[str, object]] = []

    for record in final_failures:
        records.append(
            {
                "失败类型": "远程查询最终失败",
                "配置名称": record.get("configured_name", ""),
                "证券代码": record.get("code", ""),
                "第一轮尝试次数": PRIMARY_MAX_ATTEMPTS,
                "最终统一重试次数": FINAL_RETRY_ATTEMPTS,
                "最后失败原因": record.get("reason", ""),
            }
        )

    for record in config_failures:
        records.append(
            {
                "失败类型": "本地配置错误",
                "配置名称": record.get("configured_name", ""),
                "证券代码": record.get("code", ""),
                "第一轮尝试次数": 0,
                "最终统一重试次数": 0,
                "最后失败原因": record.get("reason", ""),
            }
        )

    return pd.DataFrame(records)


def print_final_failure_alert(failure_df: pd.DataFrame) -> None:
    """使用醒目的边框和红色 ANSI 文本输出最终失败提示。"""
    if failure_df.empty:
        print("\n✓ 所有有效配置的标的均已成功完成，没有最终失败项。")
        return

    red = "\033[91m"
    bold = "\033[1m"
    reset = "\033[0m"
    border = "!" * 96

    print("\n" + red + bold + border)
    print(
        f"!!! 重要警告：共有 {len(failure_df)} 个标的最终仍未成功，"
        "它们没有写入正常指标汇总，请检查下列清单。"
    )
    print(border + reset)

    print(
        failure_df[
            ["失败类型", "配置名称", "证券代码", "最后失败原因"]
        ].to_string(index=False)
    )

    print(red + bold + border)
    print("!!! 主任务未被卡住，其余成功标的已经继续处理并正常导出。")
    print(border + reset + "\n")


def _run_main(program_start_perf: float) -> None:
    print("=" * 72)
    print("Baostock ETF 数据、远程名称校验与交叉预测工具")
    print(f"全局起始日期：{GLOBAL_START_DATE}")
    print(f"结束日期：{END_DATE}")
    print(f"本次计划处理：{len(ETF_CONFIG)} 个标的")
    print(
        f"超时策略：每次最多等待 {REQUEST_TIMEOUT_SECONDS} 秒；"
        f"第一轮每个标的最多 {PRIMARY_MAX_ATTEMPTS} 次；"
        f"第一轮结束后统一再试 {FINAL_RETRY_ATTEMPTS} 次"
    )
    print("=" * 72)

    all_data: Dict[str, pd.DataFrame] = {}
    total_count = len(ETF_CONFIG)

    # 第一轮连续尝试 3 次后仍失败的标的，统一放到这里。
    pending_final_retry: List[Dict[str, object]] = []

    # 配置格式错误不属于网络问题，不做远程重试。
    config_failures: List[Dict[str, object]] = []

    progress_bar_format = (
        "{l_bar}{bar}| {n_fmt}/{total_fmt} "
        "[已用 {elapsed} < 剩余 {remaining}, {rate_fmt}] {postfix}"
    )

    # ========================== 第一轮：逐个处理，每个最多 3 次 ==========================
    with tqdm(
        total=total_count,
        desc="第一轮总进度",
        unit="个",
        dynamic_ncols=True,
        mininterval=0.2,
        bar_format=progress_bar_format,
    ) as progress_bar:
        for index, item in enumerate(ETF_CONFIG, start=1):
            item_start_perf = time.perf_counter()

            try:
                configured_name, code = parse_config_item(item)
            except ValueError as exc:
                configured_name = str(item[0]).strip() if isinstance(item, (tuple, list)) and item else ""
                code = str(item[1]).strip() if isinstance(item, (tuple, list)) and len(item) > 1 else ""
                reason = str(exc)
                config_failures.append(
                    {
                        "configured_name": configured_name,
                        "code": code,
                        "reason": reason,
                    }
                )

                progress_bar.update(1)
                progress_bar.set_postfix_str("配置错误，已跳过")
                tqdm.write(
                    f"✗ 跑完第 {index}/{total_count} 个：配置错误｜{reason}"
                )
                continue

            df, failure_reason, attempts_used = run_etf_with_retries(
                code=code,
                configured_name=configured_name,
                max_attempts=PRIMARY_MAX_ATTEMPTS,
                phase_name=f"第一轮 {index}/{total_count}",
                progress_bar=progress_bar,
            )

            item_elapsed = time.perf_counter() - item_start_perf
            total_elapsed = time.perf_counter() - program_start_perf

            if df is not None and not df.empty:
                all_data[code] = df
                item_status = "成功"
                status_marker = "✓"
            else:
                pending_final_retry.append(
                    {
                        "configured_name": configured_name,
                        "code": code,
                        "reason": failure_reason,
                    }
                )
                item_status = "等待最终统一重试"
                status_marker = "⚠"

            progress_bar.update(1)
            progress_bar.set_postfix_str(
                f"{item_status}｜本个 {format_elapsed_time(item_elapsed)}｜"
                f"累计 {format_elapsed_time(total_elapsed)}"
            )

            tqdm.write(
                f"{status_marker} 跑完第 {index}/{total_count} 个："
                f"{configured_name}（{code}）｜状态：{item_status}｜"
                f"第一轮尝试 {attempts_used} 次｜"
                f"本个耗时：{format_elapsed_time(item_elapsed)}｜"
                f"累计耗时：{format_elapsed_time(total_elapsed)}"
            )

        progress_bar.set_description_str("第一轮处理完成")
        progress_bar.set_postfix_str(
            f"成功 {len(all_data)}｜待最终重试 {len(pending_final_retry)}｜"
            f"配置错误 {len(config_failures)}"
        )

    # ========================== 第二轮：所有第一轮失败项统一再试 1 次 ==========================
    final_failures: List[Dict[str, object]] = []

    if pending_final_retry:
        print("\n" + "=" * 96)
        print(
            f"第一轮结束：有 {len(pending_final_retry)} 个标的仍未成功，"
            f"现在统一进行最后 {FINAL_RETRY_ATTEMPTS} 次尝试。"
        )
        print("=" * 96)

        with tqdm(
            total=len(pending_final_retry),
            desc="最终统一重试",
            unit="个",
            dynamic_ncols=True,
            mininterval=0.2,
            bar_format=progress_bar_format,
        ) as retry_bar:
            for retry_index, record in enumerate(pending_final_retry, start=1):
                configured_name = str(record["configured_name"])
                code = str(record["code"])
                retry_start = time.perf_counter()

                df, failure_reason, attempts_used = run_etf_with_retries(
                    code=code,
                    configured_name=configured_name,
                    max_attempts=FINAL_RETRY_ATTEMPTS,
                    phase_name=(
                        f"最终重试 {retry_index}/{len(pending_final_retry)}"
                    ),
                    progress_bar=retry_bar,
                )

                retry_elapsed = time.perf_counter() - retry_start

                if df is not None and not df.empty:
                    all_data[code] = df
                    retry_status = "最终重试成功"
                    marker = "✓"
                else:
                    final_failures.append(
                        {
                            "configured_name": configured_name,
                            "code": code,
                            "reason": failure_reason,
                        }
                    )
                    retry_status = "最终仍失败"
                    marker = "✗"

                retry_bar.update(1)
                retry_bar.set_postfix_str(
                    f"{retry_status}｜本个 {format_elapsed_time(retry_elapsed)}"
                )
                tqdm.write(
                    f"{marker} 最终重试第 {retry_index}/"
                    f"{len(pending_final_retry)} 个："
                    f"{configured_name}（{code}）｜{retry_status}｜"
                    f"耗时 {format_elapsed_time(retry_elapsed)}"
                )

            retry_bar.set_description_str("最终统一重试完成")
            retry_bar.set_postfix_str(
                f"最终成功 {len(pending_final_retry) - len(final_failures)}｜"
                f"最终失败 {len(final_failures)}"
            )

    success_count = len(all_data)
    final_failed_count = len(final_failures) + len(config_failures)

    print("\n" + "=" * 96)
    print(
        f"远程处理全部结束：成功 {success_count}/{total_count} 个；"
        f"最终失败 {final_failed_count}/{total_count} 个"
    )
    print(
        f"行情获取与指标计算累计耗时："
        f"{format_elapsed_time(time.perf_counter() - program_start_perf)}"
    )
    print("=" * 96)

    failure_df = build_failure_dataframe(
        final_failures=final_failures,
        config_failures=config_failures,
    )

    # 先打印最终失败警告；即使存在失败项，仍继续导出所有成功结果。
    print_final_failure_alert(failure_df)

    # ========================== 构建汇总 ==========================
    if all_data:
        summary_df = pd.DataFrame(
            [build_summary_row(df) for df in all_data.values()]
        )
    else:
        summary_df = pd.DataFrame()

    if not summary_df.empty:
        display_columns = [
            "远程标的名称",
            "配置名称",
            "证券代码",
            "名称校验",
            "最新日期",
            "日MACD水上水下状态",
            "日MACD当前",
            "日MACD已持续交易日",
            "日MACD预测",
            "日KDJ当前",
            "日KDJ已持续交易日",
            "日KDJ预测",
            "周MACD水上水下状态",
            "周MACD当前",
            "周MACD已持续周数",
            "周MACD预测",
            "周KDJ当前",
            "周KDJ已持续周数",
            "周KDJ预测",
        ]

        console_summary_df = summary_df[display_columns].copy()
        marker_pairs = {
            "日MACD当前": "日MACD已持续交易日",
            "日KDJ当前": "日KDJ已持续交易日",
            "周MACD当前": "周MACD已持续周数",
            "周KDJ当前": "周KDJ已持续周数",
        }

        for status_column, duration_column in marker_pairs.items():
            console_summary_df[status_column] = [
                add_console_cross_marker(status, duration)
                for status, duration in zip(
                    summary_df[status_column],
                    summary_df[duration_column],
                )
            ]

        print("\n最终汇总（名称来自 Baostock 远程信息）：")
        print(console_summary_df.to_string(index=False))
        print(
            f"{NEW_CROSS_MARKER} 表示该日线/周线 MACD 或 KDJ 的持续时间为 1，"
            "即本期刚发生金叉或死叉。"
        )
    else:
        print("\n没有任何成功数据；仍会导出失败列表，程序不会卡住。")

    # ========================== Excel 导出 ==========================
    output_path = Path(
        f"Baostock_ETF_远程名称校验_交叉预测_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    export_start_perf = time.perf_counter()
    print(f"\n开始导出 Excel：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    used_sheet_names = set()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="汇总", index=False)
            used_sheet_names.add("汇总")

            apply_summary_cross_color_format(
                writer,
                summary_df,
                sheet_name="汇总",
            )
        else:
            pd.DataFrame(
                {
                    "提示": [
                        "本次没有任何标的成功返回，请查看“失败列表”Sheet。"
                    ]
                }
            ).to_excel(writer, sheet_name="汇总", index=False)
            used_sheet_names.add("汇总")

        # 最终失败项始终单独写入 Sheet；没有失败时写入明确提示。
        if failure_df.empty:
            pd.DataFrame(
                {"结果": ["没有最终失败项"]}
            ).to_excel(writer, sheet_name="失败列表", index=False)
        else:
            failure_df.to_excel(
                writer,
                sheet_name="失败列表",
                index=False,
            )

            failure_sheet = writer.sheets["失败列表"]
            warning_font = Font(color="FFFFFF", bold=True)
            warning_fill = PatternFill(
                fill_type="solid",
                fgColor="C00000",
            )
            for cell in failure_sheet[1]:
                cell.font = warning_font
                cell.fill = warning_fill

        used_sheet_names.add("失败列表")

        for code, df in all_data.items():
            display_name = df.attrs.get("display_name") or code
            raw_sheet_name = f"{display_name}_{code.split('.')[-1]}"
            sheet_name = make_unique_sheet_name(
                raw_sheet_name,
                used_sheet_names,
            )
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
            apply_detail_cross_color_format(
                writer,
                sheet_name=sheet_name,
            )

    export_elapsed = time.perf_counter() - export_start_perf
    print(
        f"Excel 已导出：{output_path.resolve()}｜"
        f"导出耗时：{format_elapsed_time(export_elapsed)}"
    )
    print("预测标志说明：1=快金叉，-1=快死叉，0=无临近交叉")
    print("交叉事件说明：1=本期刚金叉，-1=本期刚死叉，0=本期无新交叉")
    print("持续期数说明：交叉发生当日/当周记为 1，状态不变则逐期加 1。")
    print(
        "日/周MACD水位说明：DIF、DEA都在零轴上方=完全水上；"
        "都在零轴下方=完全水下；一上一下则显示对应过渡方向。"
    )
    print(
        "颜色说明：当日/本周刚金叉使用绿色，刚死叉使用红色；"
        "快金叉使用淡绿色，快死叉使用淡红色。"
    )
    print(
        f"控制台说明：{NEW_CROSS_MARKER} 表示该指标持续时间为 1，"
        "即本期刚发生金叉或死叉。"
    )
    print(
        "超时保护说明：每次远程处理都在独立子进程中运行；"
        "超过 30 秒会被强制结束，不会无限卡住主任务。"
    )
    print("注意：当本周尚未结束时，周线指标和下一周预测会随本周后续行情变化。")

    # Excel 导出后再次醒目提示，避免失败信息被大量汇总输出淹没。
    print_final_failure_alert(failure_df)

def run() -> None:
    program_start_datetime = datetime.now()
    program_start_perf = time.perf_counter()

    print("\n" + "=" * 72)
    print(
        f"程序开始时间："
        f"{program_start_datetime.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    print("=" * 72)

    try:
        _run_main(program_start_perf)
    finally:
        program_end_datetime = datetime.now()
        total_elapsed = time.perf_counter() - program_start_perf

        print("\n" + "=" * 72)
        print(
            f"程序结束时间："
            f"{program_end_datetime.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        print(f"全部程序总耗时：{format_elapsed_time(total_elapsed)}")
        print("=" * 72)


if __name__ == "__main__":
    mp.freeze_support()
    run()
