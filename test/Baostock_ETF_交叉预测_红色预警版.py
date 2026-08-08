import re
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import baostock as bs
import pandas as pd
from openpyxl.styles import Font, PatternFill



# ========================== 标的配置 ==========================
# 支持二元组或三元组；程序只读取前两个字段：
#   (配置名称, Baostock代码)
# 第三个字段即使存在也不会用于最终名称。
# 最终显示名称和 Excel Sheet 名称优先使用 Baostock 远程返回的证券名称。
ETF_CONFIG1 = [
    ("科创50ETF", "sh.588000"),
    ("中证银行ETF", "sh.512800"),
    # ("黄金ETF", "sh.518880"),
    # ("恒生科技ETF", "sh.513180"),
    # ("港股通科技ETF", "sz.159636"),
    # ("纳指100ETF", "sh.513100"),
    # ("新能电池ETF", "sz.159755"),
    # ("电网设备ETF", "sz.159326"),
]
ETF_CONFIG = [
    ("国证自由现金流ETF", "sz.159201"),
    ("全指红利质量ETF", "sz.159209"),
    ("航空航天ETF", "sz.159227"),
    ("恒生港股通创新药ETF", "sz.159316"),
    ("中证电网设备ETF", "sz.159326"),
    ("富时A股自由现金流ETF", "sz.159399"),
    ("半导体设备ETF", "sz.159516"),
    ("恒生高股息低波ETF", "sz.159545"),
    ("储能电池ETF", "sz.159566"),
    ("中概互联ETF", "sz.159605"),
    ("电力ETF", "sz.159611"),
    ("绿色电力ETF", "sz.159625"),
    ("港股通科技30ETF", "sz.159636"),
    ("工业母机ETF", "sz.159667"),
    ("港股通高股息ETF", "sz.159691"),
    ("消费电子ETF", "sz.159732"),
    ("沪港深云计算ETF", "sz.159738"),
    ("电池ETF", "sz.159755"),
    ("中证红利质量ETF", "sz.159758"),
    ("旅游ETF", "sz.159766"),
    ("科创创业50ETF", "sz.159781"),
    ("港股通互联网ETF", "sz.159792"),
    ("中证人工智能ETF", "sz.159819"),
    ("农业ETF", "sz.159825"),
    ("软件ETF", "sz.159852"),
    ("互联网龙头ETF", "sz.159856"),
    ("游戏ETF", "sz.159869"),
    ("化工ETF", "sz.159870"),
    ("恒生医药ETF", "sz.159892"),
    ("深证100ETF", "sz.159901"),
    ("深证红利ETF", "sz.159905"),
    ("TMT50ETF", "sz.159909"),
    ("创业板ETF", "sz.159915"),
    ("恒生ETF", "sz.159920"),
    ("消费ETF", "sz.159928"),
    ("金融地产ETF", "sz.159940"),
    ("创业板50ETF", "sz.159949"),
    ("创业板价值ETF", "sz.159966"),
    ("创业板成长ETF", "sz.159967"),
    ("有色ETF", "sz.159980"),
    ("能源化工ETF", "sz.159981"),
    ("中证创新药ETF", "sz.159992"),
    ("国证芯片ETF", "sz.159995"),
    ("家电ETF", "sz.159996"),
    ("电子ETF", "sz.159997"),
    ("计算机ETF", "sz.159998"),
    ("上证50ETF", "sh.510050"),
    ("沪深300ETF", "sh.510300"),
    ("中证500ETF", "sh.510500"),
    ("红利国企ETF", "sh.510720"),
    ("恒生国企ETF", "sh.510900"),
    ("医药ETF", "sh.512010"),
    ("证券保险ETF", "sh.512070"),
    ("中证1000ETF", "sh.512100"),
    ("中证医疗ETF", "sh.512170"),
    ("房地产ETF", "sh.512200"),
    ("生物医药ETF", "sh.512290"),
    ("有色金属ETF", "sh.512400"),
    ("半导体ETF", "sh.512480"),
    ("富时中国A50ETF", "sh.512550"),
    ("中证军工ETF", "sh.512660"),
    ("酒ETF", "sh.512690"),
    ("军工龙头ETF", "sh.512710"),
    ("CES芯片ETF", "sh.512760"),
    ("中证银行ETF", "sh.512800"),
    ("中证证券ETF", "sh.512880"),
    ("中证红利低波ETF", "sh.512890"),
    ("央企改革ETF", "sh.512950"),
    ("传媒ETF", "sh.512980"),
    ("中概互联网ETF", "sh.513050"),
    ("纳斯达克100ETF", "sh.513100"),
    ("恒生科技ETF", "sh.513180"),
    ("港股通金融ETF", "sh.513190"),
    ("中韩半导体ETF", "sh.513310"),
    ("恒生互联网ETF", "sh.513330"),
    ("港股通50ETF", "sh.513550"),
    ("标普港股低波红利ETF", "sh.513630"),
    ("标普500ETF", "sh.513650"),
    ("港股通非银ETF", "sh.513750"),
    ("中证港股通创新药ETF", "sh.513780"),
    ("日经225ETF", "sh.513880"),
    ("中证港股通央企红利ETF", "sh.513910"),
    ("恒生港股通央企红利ETF", "sh.513920"),
    ("恒生消费ETF", "sh.513970"),
    ("科技ETF", "sh.515000"),
    ("中证新能源汽车ETF", "sh.515030"),
    ("5G通信ETF", "sh.515050"),
    ("红利低波100ETF", "sh.515100"),
    ("食品饮料ETF", "sh.515170"),
    ("中证红利ETF", "sh.515180"),
    ("钢铁ETF", "sh.515210"),
    ("300红利低波ETF", "sh.515300"),
    ("大数据ETF", "sh.515400"),
    ("消费50ETF", "sh.515650"),
    ("新能源汽车产业ETF", "sh.515700"),
    ("科技50ETF", "sh.515750"),
    ("光伏ETF", "sh.515790"),
    ("通信设备ETF", "sh.515880"),
    ("稀土ETF", "sh.516150"),
    ("中证新能源ETF", "sh.516160"),
    ("云计算大数据ETF", "sh.516510"),
    ("中证智能汽车ETF", "sh.516520"),
    ("消费服务ETF", "sh.516600"),
    ("智能制造ETF", "sh.516800"),
    ("基建ETF", "sh.516970"),
    ("央企共赢ETF", "sh.517090"),
    ("中国国企ETF", "sh.517180"),
    ("沪港深消费龙头ETF", "sh.517550"),
    ("人工智能50ETF", "sh.517800"),
    ("黄金ETF", "sh.518880"),
    ("恒生创新药ETF", "sh.520500"),
    ("国新港股通央企红利ETF", "sh.520990"),
    ("MSCI中国A50ETF", "sh.560050"),
    ("央企科技ETF", "sh.560170"),
    ("碳中和ETF", "sh.561190"),
    ("恒生A股电网设备ETF", "sh.561380"),
    ("机器人ETF", "sh.562500"),
    ("稀有金属ETF", "sh.562800"),
    ("高端制造ETF", "sh.562910"),
    ("中证2000ETF", "sh.563300"),
    ("中证A500ETF", "sh.563360"),
    ("科创50ETF", "sh.588000"),
    ("科创100ETF", "sh.588030"),
    ("科创AIETF", "sh.588790"),
    ("科创综指ETF", "sh.589000"),
]

GLOBAL_START_DATE = "2021-01-05"
END_DATE = datetime.now().strftime("%Y-%m-%d")

# 预测标志：1=快金叉，-1=快死叉，0=预测下一期不会交叉
PREDICT_GOLDEN_CROSS = 1
PREDICT_DEATH_CROSS = -1
PREDICT_NO_CROSS = 0


# ========================== 基础工具 ==========================
def normalize_name(name: object) -> str:
    """名称比较时忽略空白并统一大小写。"""
    return re.sub(r"\s+", "", str(name or "")).upper()


def parse_config_item(item: Sequence[str]) -> Tuple[str, str]:
    """兼容二元组和三元组配置。"""
    if not isinstance(item, (tuple, list)) or len(item) < 2:
        raise ValueError(f"ETF_CONFIG 项格式错误：{item!r}，至少需要 (名称, 代码)")
    configured_name = str(item[0]).strip()
    code = str(item[1]).strip()
    if not configured_name or not code:
        raise ValueError(f"ETF_CONFIG 项存在空名称或空代码：{item!r}")
    return configured_name, code


def status_text(value: object) -> str:
    try:
        value = int(value)
    except (TypeError, ValueError):
        return "未知"
    if value == 1:
        return "金叉"
    if value == -1:
        return "死叉"
    return "中性"


def forecast_text(value: object) -> str:
    try:
        value = int(value)
    except (TypeError, ValueError):
        return "无法预测"
    if value == PREDICT_GOLDEN_CROSS:
        return "快金叉"
    if value == PREDICT_DEATH_CROSS:
        return "快死叉"
    return "无临近交叉"


def format_number(value: object, digits: int = 6) -> str:
    try:
        if pd.isna(value):
            return "NA"
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return "NA"


def make_unique_sheet_name(raw_name: str, used_names: set) -> str:
    """生成合法且唯一的 Excel Sheet 名称。"""
    clean = re.sub(r"[\\/*?:\[\]]", "_", str(raw_name)).strip()
    clean = clean or "Sheet"
    clean = clean[:31]

    candidate = clean
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{clean[:31 - len(suffix)]}{suffix}"
        index += 1

    used_names.add(candidate)
    return candidate


# ========================== 新交叉预警 ==========================
NEW_CROSS_MARKER = "★"
NEW_CROSS_FONT_COLOR = "C00000"       # 深红色字体
NEW_CROSS_FILL_COLOR = "FFC7CE"       # 浅红色背景


def is_first_cross_period(value: object) -> bool:
    """持续时间是否为 1；兼容 int、float、字符串和空值。"""
    try:
        if pd.isna(value):
            return False
        return int(float(value)) == 1
    except (TypeError, ValueError):
        return False


def add_console_cross_marker(status: object, age: object) -> str:
    """持续时间为 1 时，在控制台状态文字前加 ★。"""
    text = str(status)
    return f"{NEW_CROSS_MARKER}{text}" if is_first_cross_period(age) else text


def apply_summary_cross_alert_format(
    writer: pd.ExcelWriter,
    summary_df: pd.DataFrame,
    sheet_name: str = "汇总",
) -> None:
    """
    将 Excel 第一个 Sheet（汇总）中持续时间为 1 的 MACD/KDJ 项目标红。

    对每个刚发生交叉的指标，标红以下四个相关单元格：
      当前状态、交叉事件、持续时间、持续说明。
    """
    worksheet = writer.sheets[sheet_name]
    header_to_column = {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    duration_groups = {
        "日MACD已持续交易日": [
            "日MACD当前",
            "日MACD交叉事件",
            "日MACD已持续交易日",
            "日MACD持续说明",
        ],
        "日KDJ已持续交易日": [
            "日KDJ当前",
            "日KDJ交叉事件",
            "日KDJ已持续交易日",
            "日KDJ持续说明",
        ],
        "周MACD已持续周数": [
            "周MACD当前",
            "周MACD交叉事件",
            "周MACD已持续周数",
            "周MACD持续说明",
        ],
        "周KDJ已持续周数": [
            "周KDJ当前",
            "周KDJ交叉事件",
            "周KDJ已持续周数",
            "周KDJ持续说明",
        ],
    }

    alert_font = Font(color=NEW_CROSS_FONT_COLOR, bold=True)
    alert_fill = PatternFill(
        fill_type="solid",
        fgColor=NEW_CROSS_FILL_COLOR,
    )

    for dataframe_row_index, row in summary_df.iterrows():
        # 第 1 行是标题，所以 DataFrame 第 0 行对应 Excel 第 2 行。
        excel_row = int(dataframe_row_index) + 2

        for duration_column, related_columns in duration_groups.items():
            if duration_column not in summary_df.columns:
                continue
            if not is_first_cross_period(row[duration_column]):
                continue

            for column_name in related_columns:
                excel_column = header_to_column.get(column_name)
                if excel_column is None:
                    continue
                cell = worksheet.cell(row=excel_row, column=excel_column)
                cell.font = alert_font
                cell.fill = alert_fill


# ========================== 远程证券信息 ==========================
def query_remote_basic_info(code: str) -> Dict[str, Optional[str]]:
    """
    从 Baostock 远程查询证券名称和上市日期。

    返回字段：
      remote_code, remote_name, ipo_date, error
    """
    result: Dict[str, Optional[str]] = {
        "remote_code": code,
        "remote_name": None,
        "ipo_date": None,
        "error": None,
    }

    rs = bs.query_stock_basic(code=code)
    if rs.error_code != "0":
        result["error"] = f"{rs.error_code}: {rs.error_msg}"
        return result

    if not rs.next():
        result["error"] = "远程未返回证券基本信息"
        return result

    fields = list(getattr(rs, "fields", []) or [])
    row_data = rs.get_row_data()
    row = dict(zip(fields, row_data))

    # Baostock 官方字段通常为 code、code_name、ipoDate 等。
    remote_code = (row.get("code") or code).strip()
    remote_name = (row.get("code_name") or row.get("name") or "").strip()
    ipo_raw = (row.get("ipoDate") or "").strip()

    result["remote_code"] = remote_code or code
    result["remote_name"] = remote_name or None

    if ipo_raw:
        ipo_date = pd.to_datetime(ipo_raw, errors="coerce")
        if not pd.isna(ipo_date):
            result["ipo_date"] = ipo_date.strftime("%Y-%m-%d")

    return result


# ========================== 技术指标 ==========================
def calculate_macd(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    ema_fast = df["close"].ewm(span=10, adjust=False).mean()
    ema_slow = df["close"].ewm(span=22, adjust=False).mean()
    dif = ema_fast - ema_slow
    dea = dif.ewm(span=9, adjust=False).mean()
    macd = (dif - dea) * 2

    df[f"{prefix}dif"] = dif
    df[f"{prefix}dea"] = dea
    df[f"{prefix}macd"] = macd
    return df


def calculate_kdj(df: pd.DataFrame, prefix: str = "") -> pd.DataFrame:
    low_list = df["low"].rolling(9, min_periods=1).min()
    high_list = df["high"].rolling(9, min_periods=1).max()

    hl_range = (high_list - low_list).mask((high_list - low_list) == 0)
    rsv = (df["close"] - low_list) / hl_range * 100
    rsv = rsv.fillna(50).ffill()

    k = rsv.ewm(com=2, adjust=False).mean()
    d = k.ewm(com=2, adjust=False).mean()
    j = 3 * k - 2 * d

    df[f"{prefix}k"] = k
    df[f"{prefix}d"] = d
    df[f"{prefix}j"] = j
    return df


def add_cross_status_metrics(
    df: pd.DataFrame,
    left_col: str,
    right_col: str,
    prefix: str,
    age_col: str,
    period_label: str,
) -> pd.DataFrame:
    """
    计算金叉/死叉当前状态、交叉事件和当前状态持续期数。

    生成字段：
      {prefix}_status
        1  = 左线在右线上方，当前处于金叉状态
        -1 = 左线在右线下方，当前处于死叉状态
        0  = 两线始终相等，无法判断

      {prefix}_cross_event
        1  = 本期刚刚由死叉转为金叉
        -1 = 本期刚刚由金叉转为死叉
        0  = 本期没有发生新的交叉

      {age_col}
        当前金叉/死叉状态已经连续保持的期数。
        交叉发生当期为 1，下一期若状态不变则为 2，以此类推。

      {prefix}_cross_age_text
        例如“金叉第3个交易日”“死叉第2周”。

    注意：如果数据起点之前已经发生交叉，数据第一段状态的持续期数只能
    从当前数据集起点开始计算；后续发生的交叉可以准确从 1 开始计数。
    """
    status_col = f"{prefix}_status"
    event_col = f"{prefix}_cross_event"
    event_text_col = f"{prefix}_cross_event_text"
    age_text_col = f"{prefix}_cross_age_text"

    gap = df[left_col] - df[right_col]

    status = pd.Series(pd.NA, index=df.index, dtype="Int64")
    status.loc[gap > 0] = 1
    status.loc[gap < 0] = -1

    # 两线暂时相等时沿用最近状态；开头连续相等时使用后续第一个有效状态。
    status = status.ffill().bfill().fillna(0).astype(int)
    df[status_col] = status

    previous_status = status.shift(1)
    event = pd.Series(0, index=df.index, dtype="int64")
    event.loc[(status == 1) & (previous_status == -1)] = 1
    event.loc[(status == -1) & (previous_status == 1)] = -1
    df[event_col] = event

    if period_label == "周":
        event_text_map = {1: "本周金叉", -1: "本周死叉", 0: "本周无新交叉"}
    else:
        event_text_map = {1: "当日金叉", -1: "当日死叉", 0: "当日无新交叉"}
    df[event_text_col] = df[event_col].map(event_text_map)

    # 每当状态变化就开启一个新分组；组内序号从 1 开始。
    run_id = status.ne(status.shift(1)).cumsum()
    age = status.groupby(run_id).cumcount() + 1
    age = age.where(status != 0, 0).astype(int)
    df[age_col] = age

    age_text = []
    for current_status, current_age in zip(status.tolist(), age.tolist()):
        if current_status == 1:
            age_text.append(f"金叉第{current_age}{period_label}")
        elif current_status == -1:
            age_text.append(f"死叉第{current_age}{period_label}")
        else:
            age_text.append("中性")
    df[age_text_col] = age_text

    return df


def add_linear_cross_forecast(
    df: pd.DataFrame,
    left_col: str,
    right_col: str,
    prefix: str,
) -> pd.DataFrame:
    """
    使用最近两期“线差”做一次线性外推。

    gap_t      = left_t - right_t
    next_gap   = gap_t + (gap_t - gap_t-1)
               = 2 * gap_t - gap_t-1

    判断：
      当前 gap <= 0 且预测 next_gap > 0  -> 快金叉，flag=1
      当前 gap >= 0 且预测 next_gap < 0  -> 快死叉，flag=-1
      否则                              -> 无临近交叉，flag=0
    """
    gap_col = f"{prefix}_gap"
    previous_gap_col = f"{prefix}_prev_gap"
    predicted_gap_col = f"{prefix}_next_gap_pred"
    flag_col = f"{prefix}_near_cross_flag"
    text_col = f"{prefix}_near_cross_text"

    df[gap_col] = df[left_col] - df[right_col]
    df[previous_gap_col] = df[gap_col].shift(1)
    df[predicted_gap_col] = 2 * df[gap_col] - df[previous_gap_col]

    flag = pd.Series(PREDICT_NO_CROSS, index=df.index, dtype="int64")
    valid = df[previous_gap_col].notna() & df[predicted_gap_col].notna()

    golden = valid & (df[gap_col] <= 0) & (df[predicted_gap_col] > 0)
    death = valid & (df[gap_col] >= 0) & (df[predicted_gap_col] < 0)

    flag.loc[golden] = PREDICT_GOLDEN_CROSS
    flag.loc[death] = PREDICT_DEATH_CROSS

    df[flag_col] = flag
    df[text_col] = df[flag_col].map(
        {
            PREDICT_GOLDEN_CROSS: "快金叉",
            PREDICT_DEATH_CROSS: "快死叉",
            PREDICT_NO_CROSS: "无临近交叉",
        }
    )
    return df


# ========================== 周线生成与预测 ==========================
def add_weekly_indicators(daily_df: pd.DataFrame) -> pd.DataFrame:
    df = daily_df.copy()

    if not pd.api.types.is_datetime64_any_dtype(df["date"]):
        print("  检测到 date 不是 datetime，正在转换...")
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).reset_index(drop=True)
        if df.empty:
            raise ValueError("所有日期转换失败，无法生成周线")

    df["date"] = df["date"].dt.normalize()

    weekly_df = (
        df.resample("W-FRI", on="date")
        .agg(
            {
                "open": "first",
                "high": "max",
                "low": "min",
                "close": "last",
                "volume": "sum",
                "amount": "sum",
            }
        )
        .reset_index()
    )

    # 去掉整周无交易造成的空周，避免技术指标被空值污染。
    weekly_df = weekly_df.dropna(subset=["close"]).reset_index(drop=True)
    weekly_df["date"] = weekly_df["date"].dt.normalize()

    weekly_df = calculate_macd(weekly_df, prefix="week_")
    weekly_df = calculate_kdj(weekly_df, prefix="week_")

    # 周线状态、交叉事件和持续周数必须在“唯一周线数据”上计算，
    # 不能在重复映射后的日线上计算，否则持续周数会被交易日重复放大。
    weekly_df = add_cross_status_metrics(
        weekly_df,
        left_col="week_dif",
        right_col="week_dea",
        prefix="week_macd",
        age_col="week_macd_cross_weeks",
        period_label="周",
    )
    weekly_df = add_cross_status_metrics(
        weekly_df,
        left_col="week_k",
        right_col="week_d",
        prefix="week_kdj",
        age_col="week_kdj_cross_weeks",
        period_label="周",
    )

    # 周 MACD 和周 KDJ：使用最近两个周线差值预测下一周。
    weekly_df = add_linear_cross_forecast(
        weekly_df, "week_dif", "week_dea", "week_macd"
    )
    weekly_df = add_linear_cross_forecast(
        weekly_df, "week_k", "week_d", "week_kdj"
    )

    df["week_end_date"] = (
        df["date"].dt.to_period("W-FRI").dt.end_time.dt.normalize()
    )

    weekly_merge = weekly_df.rename(columns={"date": "week_end_date"}).copy()
    weekly_columns = [
        "week_end_date",
        "week_dif",
        "week_dea",
        "week_macd",
        "week_k",
        "week_d",
        "week_j",
        "week_macd_status",
        "week_macd_cross_event",
        "week_macd_cross_event_text",
        "week_macd_cross_weeks",
        "week_macd_cross_age_text",
        "week_kdj_status",
        "week_kdj_cross_event",
        "week_kdj_cross_event_text",
        "week_kdj_cross_weeks",
        "week_kdj_cross_age_text",
        "week_macd_gap",
        "week_macd_prev_gap",
        "week_macd_next_gap_pred",
        "week_macd_near_cross_flag",
        "week_macd_near_cross_text",
        "week_kdj_gap",
        "week_kdj_prev_gap",
        "week_kdj_next_gap_pred",
        "week_kdj_near_cross_flag",
        "week_kdj_near_cross_text",
    ]

    df = pd.merge(
        df,
        weekly_merge[weekly_columns],
        on="week_end_date",
        how="left",
        validate="many_to_one",
    )
    df = df.drop(columns=["week_end_date"])

    matched_count = int(df["week_dif"].notna().sum())
    print(f"  周线指标匹配成功：{matched_count}/{len(df)} 行")
    return df


# ========================== 单标的数据获取 ==========================
def get_etf_full_data(code: str, configured_name: str) -> Optional[pd.DataFrame]:
    print(f"\n{'=' * 72}")
    print(f"配置名称：{configured_name}")
    print(f"证券代码：{code}")
    print("-" * 72)

    login_result = bs.login()
    if login_result.error_code != "0":
        print(f"Baostock 登录失败：{login_result.error_msg}")
        return None

    try:
        basic = query_remote_basic_info(code)
        remote_name = basic.get("remote_name")
        remote_code = basic.get("remote_code") or code
        ipo_date = basic.get("ipo_date")

        if remote_name:
            name_match = normalize_name(configured_name) == normalize_name(remote_name)
            display_name = remote_name
            print(f"远程名称：{remote_name}")
            if name_match:
                print("名称校验：一致")
            else:
                print(
                    f"名称校验：不一致。配置名称={configured_name}，"
                    f"远程名称={remote_name}；后续自动使用远程名称。"
                )
        else:
            name_match = False
            display_name = configured_name
            print(
                f"远程名称：获取失败（{basic.get('error') or '未知原因'}）；"
                "后续暂用配置名称。"
            )

        if remote_code != code:
            print(f"代码校验：远程返回代码 {remote_code}，与配置代码 {code} 不一致")
        else:
            print("代码校验：一致")

        start_date = GLOBAL_START_DATE
        if ipo_date:
            start_date = max(GLOBAL_START_DATE, ipo_date)
            print(f"上市日期：{ipo_date}，实际起始日期：{start_date}")
        else:
            print(f"上市日期：未获取，使用全局起始日期 {GLOBAL_START_DATE}")

        print(f"请求数据时间范围：{start_date} 至 {END_DATE}")

        rs = bs.query_history_k_data_plus(
            code=code,
            fields="date,open,high,low,close,volume,amount,pctChg,turn",
            start_date=start_date,
            end_date=END_DATE,
            frequency="d",
            adjustflag="2",
        )

        if rs.error_code != "0":
            print(f"获取行情数据失败：{rs.error_msg}")
            return None

        data_list: List[List[str]] = []
        while rs.next():
            data_list.append(rs.get_row_data())

        if not data_list:
            print("未获取到任何行情数据")
            return None

        df = pd.DataFrame(data_list, columns=rs.fields)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"]).reset_index(drop=True)

        if df.empty:
            print("所有日期转换失败")
            return None

        numeric_columns = [col for col in df.columns if col != "date"]
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

        # 核心行情列缺失的行不能参与指标计算。
        df = df.dropna(subset=["open", "high", "low", "close"]).reset_index(drop=True)
        if df.empty:
            print("有效 OHLC 数据为空")
            return None

        print(f"原始数据获取成功：共 {len(df)} 个交易日")
        print(f"数据起始日期：{df['date'].iloc[0].strftime('%Y-%m-%d')}")
        print(f"数据结束日期：{df['date'].iloc[-1].strftime('%Y-%m-%d')}")

        print("正在计算日线 MACD/KDJ...")
        df = calculate_macd(df, prefix="day_")
        df = calculate_kdj(df, prefix="day_")

        df = add_cross_status_metrics(
            df,
            left_col="day_dif",
            right_col="day_dea",
            prefix="day_macd",
            age_col="day_macd_cross_days",
            period_label="个交易日",
        )
        df = add_cross_status_metrics(
            df,
            left_col="day_k",
            right_col="day_d",
            prefix="day_kdj",
            age_col="day_kdj_cross_days",
            period_label="个交易日",
        )

        # 日 MACD、日 KDJ：使用最近两个交易日的线差预测下一交易日。
        # 例如上期差=1、本期差=3，则预测下期差=3+(3-1)=5。
        df = add_linear_cross_forecast(df, "day_dif", "day_dea", "day_macd")
        df = add_linear_cross_forecast(df, "day_k", "day_d", "day_kdj")

        print("正在计算周线 MACD/KDJ，并预测下一周交叉...")
        df = add_weekly_indicators(df)

        # 把远程名称校验信息写入每一行，便于 Excel 中直接核对。
        df.insert(0, "name_match", bool(remote_name) and name_match)
        df.insert(0, "remote_name", remote_name or "")
        df.insert(0, "configured_name", configured_name)
        df.insert(0, "code", code)

        df.attrs["code"] = code
        df.attrs["configured_name"] = configured_name
        df.attrs["remote_name"] = remote_name or ""
        df.attrs["display_name"] = display_name
        df.attrs["name_match"] = bool(remote_name) and name_match
        df.attrs["name_verified"] = bool(remote_name)

        latest = df.iloc[-1]
        rate = (latest["close"] / df["close"].iloc[0] - 1) * 100

        print(f"\n{display_name}（{code}）处理完成")
        print(f"总数据条数：{len(df)}")
        print(f"累计涨跌幅：{rate:.2f}%")
        print(f"最新日期：{latest['date'].strftime('%Y-%m-%d')}")
        print(
            f"日线 MACD 当前：{status_text(latest['day_macd_status'])}，"
            f"已持续 {int(latest['day_macd_cross_days'])} 个交易日"
        )
        print(
            f"日线 KDJ 当前：{status_text(latest['day_kdj_status'])}，"
            f"已持续 {int(latest['day_kdj_cross_days'])} 个交易日"
        )
        print(
            f"周线 MACD 当前：{status_text(latest['week_macd_status'])}，"
            f"已持续 {int(latest['week_macd_cross_weeks'])} 周"
        )
        print(
            f"周线 KDJ 当前：{status_text(latest['week_kdj_status'])}，"
            f"已持续 {int(latest['week_kdj_cross_weeks'])} 周"
        )

        print("\n线性外推结果：")
        print(
            "  日 MACD（预测下一交易日）："
            f"上期差={format_number(latest['day_macd_prev_gap'])}，"
            f"本期差={format_number(latest['day_macd_gap'])}，"
            f"预测下期差={format_number(latest['day_macd_next_gap_pred'])}，"
            f"标志={latest['day_macd_near_cross_text']}"
        )
        print(
            "  日 KDJ（预测下一交易日）："
            f"上期差={format_number(latest['day_kdj_prev_gap'])}，"
            f"本期差={format_number(latest['day_kdj_gap'])}，"
            f"预测下期差={format_number(latest['day_kdj_next_gap_pred'])}，"
            f"标志={latest['day_kdj_near_cross_text']}"
        )
        print(
            "  周 KDJ（预测下一周）："
            f"上周差={format_number(latest['week_kdj_prev_gap'])}，"
            f"本周差={format_number(latest['week_kdj_gap'])}，"
            f"预测下周差={format_number(latest['week_kdj_next_gap_pred'])}，"
            f"标志={latest['week_kdj_near_cross_text']}"
        )
        print(
            "  周 MACD（预测下一周）："
            f"上周差={format_number(latest['week_macd_prev_gap'])}，"
            f"本周差={format_number(latest['week_macd_gap'])}，"
            f"预测下周差={format_number(latest['week_macd_next_gap_pred'])}，"
            f"标志={latest['week_macd_near_cross_text']}"
        )

        return df

    except Exception as exc:
        print(f"处理失败：{exc}")
        traceback.print_exc()
        return None

    finally:
        try:
            bs.logout()
        except Exception:
            pass


# ========================== 汇总与导出 ==========================
def build_summary_row(df: pd.DataFrame) -> Dict[str, object]:
    latest = df.iloc[-1]
    remote_name = df.attrs.get("remote_name", "")
    configured_name = df.attrs.get("configured_name", "")
    display_name = df.attrs.get("display_name", remote_name or configured_name)
    name_verified = bool(df.attrs.get("name_verified", False))
    name_match = bool(df.attrs.get("name_match", False))

    if not name_verified:
        name_check = "远程名称未获取"
    elif name_match:
        name_check = "一致"
    else:
        name_check = "不一致，已使用远程名称"

    return {
        "远程标的名称": display_name,
        "配置名称": configured_name,
        "证券代码": df.attrs.get("code", ""),
        "名称校验": name_check,
        "起始日期": df["date"].iloc[0].strftime("%Y-%m-%d"),
        "最新日期": latest["date"].strftime("%Y-%m-%d"),
        "数据条数": len(df),
        "日MACD当前": status_text(latest["day_macd_status"]),
        "日MACD交叉事件": latest["day_macd_cross_event_text"],
        "日MACD已持续交易日": int(latest["day_macd_cross_days"]),
        "日MACD持续说明": latest["day_macd_cross_age_text"],
        "日MACD预测": latest["day_macd_near_cross_text"],
        "日MACD预测标志": int(latest["day_macd_near_cross_flag"]),
        "日MACD上期差": latest["day_macd_prev_gap"],
        "日MACD本期差": latest["day_macd_gap"],
        "日MACD预测下期差": latest["day_macd_next_gap_pred"],
        "日KDJ当前": status_text(latest["day_kdj_status"]),
        "日KDJ交叉事件": latest["day_kdj_cross_event_text"],
        "日KDJ已持续交易日": int(latest["day_kdj_cross_days"]),
        "日KDJ持续说明": latest["day_kdj_cross_age_text"],
        "日KDJ预测": latest["day_kdj_near_cross_text"],
        "日KDJ预测标志": int(latest["day_kdj_near_cross_flag"]),
        "日KDJ上期差": latest["day_kdj_prev_gap"],
        "日KDJ本期差": latest["day_kdj_gap"],
        "日KDJ预测下期差": latest["day_kdj_next_gap_pred"],
        "周KDJ当前": status_text(latest["week_kdj_status"]),
        "周KDJ交叉事件": latest["week_kdj_cross_event_text"],
        "周KDJ已持续周数": int(latest["week_kdj_cross_weeks"]),
        "周KDJ持续说明": latest["week_kdj_cross_age_text"],
        "周KDJ预测": latest["week_kdj_near_cross_text"],
        "周KDJ预测标志": int(latest["week_kdj_near_cross_flag"]),
        "周KDJ上周差": latest["week_kdj_prev_gap"],
        "周KDJ本周差": latest["week_kdj_gap"],
        "周KDJ预测下周差": latest["week_kdj_next_gap_pred"],
        "周MACD当前": status_text(latest["week_macd_status"]),
        "周MACD交叉事件": latest["week_macd_cross_event_text"],
        "周MACD已持续周数": int(latest["week_macd_cross_weeks"]),
        "周MACD持续说明": latest["week_macd_cross_age_text"],
        "周MACD预测": latest["week_macd_near_cross_text"],
        "周MACD预测标志": int(latest["week_macd_near_cross_flag"]),
        "周MACD上周差": latest["week_macd_prev_gap"],
        "周MACD本周差": latest["week_macd_gap"],
        "周MACD预测下周差": latest["week_macd_next_gap_pred"],
    }


def run() -> None:
    print("=" * 72)
    print("Baostock ETF 数据、远程名称校验与交叉预测工具")
    print(f"全局起始日期：{GLOBAL_START_DATE}")
    print(f"结束日期：{END_DATE}")
    print("=" * 72)

    all_data: Dict[str, pd.DataFrame] = {}
    success_count = 0

    for item in ETF_CONFIG:
        try:
            configured_name, code = parse_config_item(item)
        except ValueError as exc:
            print(f"跳过错误配置：{exc}")
            continue

        df = get_etf_full_data(code, configured_name)
        if df is not None and not df.empty:
            # 使用代码作为 key，避免名称相同或错误导致覆盖。
            all_data[code] = df
            success_count += 1

    print("\n" + "=" * 72)
    print(f"全部处理完成：成功 {success_count}/{len(ETF_CONFIG)} 个标的")
    print("=" * 72)

    if not all_data:
        print("未获取到任何有效数据")
        return

    summary_df = pd.DataFrame([build_summary_row(df) for df in all_data.values()])

    # 控制台最终汇总：远程名称直接放在第一列。
    display_columns = [
        "远程标的名称",
        "配置名称",
        "证券代码",
        "名称校验",
        "最新日期",
        "日MACD当前",
        "日MACD已持续交易日",
        "日MACD预测",
        "日KDJ当前",
        "日KDJ已持续交易日",
        "日KDJ预测",
        "周MACD当前",
        "周MACD已持续周数",
        "周MACD预测",
        "周KDJ当前",
        "周KDJ已持续周数",
        "周KDJ预测",
    ]
    # 控制台显示副本：持续时间为 1 的指标，在当前状态前添加 ★。
    console_summary_df = summary_df[display_columns].copy()
    marker_pairs = {
        "日MACD当前": "日MACD已持续交易日",
        "日KDJ当前": "日KDJ已持续交易日",
        "周MACD当前": "周MACD已持续周数",
        "周KDJ当前": "周KDJ已持续周数",
    }
    for status_column, duration_column in marker_pairs.items():
        console_summary_df[status_column] = [
            add_console_cross_marker(status, duration)
            for status, duration in zip(
                summary_df[status_column],
                summary_df[duration_column],
            )
        ]

    print("\n最终汇总（名称来自 Baostock 远程信息）：")
    print(console_summary_df.to_string(index=False))
    print(f"{NEW_CROSS_MARKER} 表示该日线/周线 MACD 或 KDJ 的持续时间为 1，即本期刚发生金叉或死叉。")

    output_path = Path(
        f"Baostock_ETF_远程名称校验_交叉预测_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    used_sheet_names = set()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总", index=False)
        used_sheet_names.add("汇总")

        # Sheet1 即“汇总”：持续时间为 1 的日/周 MACD、KDJ 相关单元格标红。
        apply_summary_cross_alert_format(writer, summary_df, sheet_name="汇总")

        for code, df in all_data.items():
            display_name = df.attrs.get("display_name") or code
            raw_sheet_name = f"{display_name}_{code.split('.')[-1]}"
            sheet_name = make_unique_sheet_name(raw_sheet_name, used_sheet_names)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"\nExcel 已导出：{output_path.resolve()}")
    print("预测标志说明：1=快金叉，-1=快死叉，0=无临近交叉")
    print("交叉事件说明：1=本期刚金叉，-1=本期刚死叉，0=本期无新交叉")
    print("持续期数说明：交叉发生当日/当周记为 1，状态不变则逐期加 1。")
    print(f"红色预警说明：Excel 汇总 Sheet 中持续时间为 1 的指标已标红；控制台使用 {NEW_CROSS_MARKER} 标记。")
    print("注意：当本周尚未结束时，周线指标和下一周预测会随本周后续行情变化。")


if __name__ == "__main__":
    run()
